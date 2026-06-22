from openpyxl import load_workbook

wb = load_workbook('MEDIT_Delivery Note_template.xlsx')
ws = wb.active

print("--- Template row content ---")
for r in range(1, 35):
    print(f"Row {r:02d}: {[ws.cell(row=r, column=c).value for c in range(1, 6)]}")
