from openpyxl import load_workbook

wb = load_workbook('MEDIT_Delivery Note.xlsx')
ws = wb.active

print(f"Max row in generated file: {ws.max_row}")
print("\n--- Row Heights and content in generated file ---")
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    height = ws.row_dimensions[row].height
    print(f"Row {row:02d} (Height: {height}): {val}")

print("\n--- Row Breaks ---")
for brk in ws.row_breaks.brk:
    print(f"Break at row: {brk.id}")
