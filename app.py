from concurrent.futures import ThreadPoolExecutor
import os
import re
import json
import pandas as pd
from pypdf import PdfReader, PdfWriter
from flask import Flask, render_template, request, send_file, jsonify
from io import BytesIO
import tempfile
import pypdfium2 as pdfium
import zxingcpp
from PIL import Image
import zipfile
import threading
from reportlab.pdfgen import canvas as rl_canvas

app = Flask(__name__)

# Lock for thread-safe
pdfium_lock = threading.Lock() 

# Helper to extract images from template if they don't exist
def ensure_images_extracted():
    template_path = 'MEDIT_Delivery Note_template.xlsx'
    if not os.path.exists(template_path):
        return
    
    with zipfile.ZipFile(template_path, 'r') as z:
        media_map = {
            'xl/media/image1.png': 'image1.png',
            'xl/media/image2.png': 'image2.png'
        }
        for zip_p, local_p in media_map.items():
            if not os.path.exists(local_p):
                try:
                    with open(local_p, 'wb') as f:
                        f.write(z.read(zip_p))
                except:
                    pass

ensure_images_extracted()

MAX_WORKERS = 2

def get_pdf_size(pdf_stream):
    try:
        pdf_stream.seek(0)
        r = PdfReader(pdf_stream)
        page = r.pages[0]
        w, h = float(page.mediabox.width), float(page.mediabox.height)
        # Standard A4 is ~595x842 pts. Labels are ~280x560.
        # Increasing threshold to 600 to avoid misclassifying labels as A4.
        if w > 600 or h > 600:
            return "a4"
        return "label"
    except:
        return "error"

def expand_ref_numbers(ref_str):
    if not ref_str or ref_str == "Not Found":
        return ["Not Found"]
    parts = [p.strip() for p in re.split(r'[;,]', ref_str)]
    expanded = []
    base_prefix = ""
    for p in parts:
        if not p: continue
        clean_p = re.sub(r'^RMA-?', '', p, flags=re.IGNORECASE).strip()
        digits_only = re.sub(r'\D', '', clean_p)
        if clean_p.startswith('400'):
            expanded.append(clean_p)
            base_prefix = clean_p[:6]
        elif base_prefix and 0 < len(digits_only) < 6:
            # 하이픈이 OCR로 소실돼도(예: "-4218" -> "4218") suffix로 처리
            expanded.append(base_prefix + digits_only)
        else:
            if digits_only:
                expanded.append(digits_only)
    return expanded

def decode_top_barcode(pdf_bytes, scale=3.0):

    rotations = [0, 90]

    tracking_candidates = []
    barcode_pages = None

    try:
        for rotation in rotations:

            with pdfium_lock:
                doc = pdfium.PdfDocument(
                    BytesIO(pdf_bytes)
                )

                page = doc[0]

                bitmap = page.render(
                    scale=scale,
                    rotation=rotation
                )

                img = bitmap.to_pil()

                doc.close()

            results = zxingcpp.read_barcodes(img)

            for result in results:
                text = result.text or ""

                if (
                    text.isdigit()
                    and len(text) >= 10
                ):
                    tracking_candidates.append(text)

                page_match = re.search(
                    r'\b\d+\s*/\s*(\d+)\b',
                    text
                )

                if not page_match:
                    page_match = re.search(
                        r'\b\d+\s+of\s+(\d+)\b',
                        text,
                        re.IGNORECASE
                    )

                if page_match:
                    barcode_pages = int(
                        page_match.group(1)
                    )


            if tracking_candidates and barcode_pages:
                break

        tracking_no = None

        if tracking_candidates:
            tracking_no = max(
                tracking_candidates,
                key=len
            )

        return tracking_no, barcode_pages

    except Exception as e:
        print(
            f"Barcode decode error: {e}"
        )

        return None, None

def classify_carrier(text):
    """라벨 텍스트로 DHL/FedEx 구분"""
    if re.search(r'WAYBILL', text, re.IGNORECASE):
        return 'dhl'
    if re.search(r'TRK#|MPS#|##\s*MASTER\s*##', text, re.IGNORECASE):
        return 'fedex'
    return 'dhl'  # 기본값

CATEGORY_ORDER = {
    ('dhl', False): 0,    # normal-dhl
    ('fedex', False): 1,  # normal-fedex
    ('dhl', True): 2,     # rma-dhl
    ('fedex', True): 3,   # rma-fedex
}

def category_label(carrier, is_rma):
    return ('rma-' if is_rma else 'normal-') + carrier

def decode_all_page_barcodes(pdf_bytes, total_pages, scale=3.0):

    page_results = []

    try:
        with pdfium_lock:
            doc = pdfium.PdfDocument(BytesIO(pdf_bytes))

            for page_idx in range(total_pages):
                try:
                    page = doc[page_idx]

                    bitmap = page.render(
                        scale=scale,
                        rotation=0
                    )

                    img = bitmap.to_pil()

                    results = zxingcpp.read_barcodes(img)

                    for result in results:
                        text = result.text or ""

                        # FedEx: 1/2
                        match = re.search(
                            r'\b(\d+)\s*/\s*(\d+)\b',
                            text
                        )

                        # FedEx: 1 of 2
                        if not match:
                            match = re.search(
                                r'\b(\d+)\s+of\s+(\d+)\b',
                                text,
                                re.IGNORECASE
                            )

                        if match:
                            page_results.append({
                                "page_idx": page_idx,
                                "piece_no": int(match.group(1)),
                                "total": int(match.group(2))
                            })
                            break

                except Exception as e:
                    print(
                        f"Barcode scan failed "
                        f"(page={page_idx}): {e}"
                    )

            doc.close()

    except Exception as e:
        print(
            f"Full PDF barcode scan error: {e}"
        )

    return page_results


def extract_order_no_from_filename(filename):

    stem = os.path.splitext(
        os.path.basename(filename)
    )[0]

    # RMA 접두사 제거
    stem = re.sub(
        r'^RMA-?',
        '',
        stem,
        flags=re.IGNORECASE
    )

    # 숫자와 하이픈 이외의 문자는 구분자로 변경
    candidate = re.sub(
        r'[^0-9\-]+',
        ';',
        stem
    ).strip(';')

    if not candidate:
        return None

    first_part = candidate.split(';')[0]

    # SAP Order No는 400으로 시작
    if (
        first_part.startswith('400')
        and len(first_part) >= 7
    ):
        return candidate

    return None

def extract_single_pdf(file_content, filename, manual_ref=None):
 
    try:
        # PDF 전체 페이지 수
        stream = BytesIO(file_content)
        r = PdfReader(stream)
        total_pages = len(r.pages)

        if total_pages == 0:
            return {
                "order_no": "Not Found",
                "tracking_no": "Not Found",
                "size": "error",
                "total_pages": 0,
                "barcode_pages": None
            }

        # --------------------------------------------------
        # Order No
        # --------------------------------------------------
        if manual_ref:
            order_no = manual_ref
        else:
            stem = os.path.splitext(
                os.path.basename(filename)
            )[0]

            stem = re.sub(
                r'^RMA-?',
                '',
                stem,
                flags=re.IGNORECASE
            )

            candidate = re.sub(
                r'[^0-9\-]+',
                ';',
                stem
            ).strip(';')

            order_no = candidate or "Not Found"

        # --------------------------------------------------
        # PDF Size
        # --------------------------------------------------
        size = get_pdf_size(BytesIO(file_content))

        # --------------------------------------------------
        # 첫 페이지 바코드
        # --------------------------------------------------
        tracking_no, barcode_pages = decode_top_barcode(
            file_content
        )

        return {
            "order_no": order_no,
            "tracking_no": tracking_no or "Not Found",
            "size": size,
            "total_pages": total_pages,
            "barcode_pages": barcode_pages
        }

    except Exception as e:
        print(f"Error processing {filename}: {e}")

        return {
            "order_no": manual_ref or "Not Found",
            "tracking_no": "Error",
            "size": "error",
            "total_pages": 0,
            "barcode_pages": None
        }

@app.route('/')
def index():
    return render_template('index.html')

progress_data = {}

@app.route('/progress/<job_id>')
def get_progress(job_id):
    return jsonify(progress_data.get(job_id, {"current": 0, "total": 0}))

@app.route('/parse', methods=['POST'])
def parse():
    job_id = request.form.get('job_id')
    files = request.files.getlist('files')
    manual_refs_json = request.form.get('manual_refs', '{}')
    try:
        manual_refs = json.loads(manual_refs_json)
    except:
        manual_refs = {}
    
    if job_id:
        progress_data[job_id] = {"current": 0, "total": len(files)}

    file_data = []
    for f in files:
        content = f.read()
        m_ref = manual_refs.get(f.filename)
        file_data.append((content, f.filename, m_ref))

    results = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [
            executor.submit(
                extract_single_pdf,
                d[0],
                d[1],
                d[2]
            )
            for d in file_data
        ]

        for future in futures:
            result = future.result()

            expanded_orders = expand_ref_numbers(
                result["order_no"]
            )

            for ref in expanded_orders:
                results.append({
                    "Ref No": ref,
                    "Tracking Number": result["tracking_no"],
                    "Size Type": result["size"],
                    "Total Pages": result["total_pages"],
                    "Barcode Pages": result["barcode_pages"]
                })

            if job_id and job_id in progress_data:
                progress_data[job_id]["current"] += 1

    if job_id and job_id in progress_data:
        del progress_data[job_id]

    return jsonify(results)

from datetime import datetime

@app.route('/parse_picking_list', methods=['POST'])
def parse_picking_list_endpoint():
    file = request.files.get('picking_list')
    tracking_data_json = request.form.get('tracking_data', '[]')
    missing_overrides_json = request.form.get('missing_overrides', '{}')  # 추가
    if not file:
        return jsonify({"error": "Picking List 파일을 업로드해주세요."}), 400
    try:
        tracking_map = {}
        sap_orders = set()
        try:
            tracking_list = json.loads(tracking_data_json)
            for item in tracking_list:
                ref = item.get('Ref No', '').strip()
                track = item.get('Tracking Number', '').strip()
                if ref and ref != "Not Found":
                    sap_orders.add(ref)
                if ref and track and track != "Not Found":
                    tracking_map[ref] = track
        except:
            pass

        # missing_overrides: { "4000003968": "20260708_4000003968_MF" }
        try:
            missing_overrides = json.loads(missing_overrides_json)
        except:
            missing_overrides = {}

        stream = BytesIO(file.read())
        reader = PdfReader(stream)

        try:
            start_page = int(request.form.get('start_page', 1)) - 1
            if start_page < 0: start_page = 0
        except:
            start_page = 0

        orders = []
        seen_packing_nos = set()
        matched_orders = set()  # Picking List에서 매칭된 SAP Order 추적

        for i in range(start_page, len(reader.pages)):
            text = reader.pages[i].extract_text()
            if not text or ("Packing No." not in text or "OrderNo." not in text):
                continue
            parts = re.split(r'(Packing No\.)', text)
            for j in range(1, len(parts), 2):
                if j+1 < len(parts):
                    segment = parts[j] + parts[j+1]
                    p_match = re.search(r'Packing No\.\s*(\S+)', segment)
                    o_match = re.search(r'OrderNo\.\s*(\S+)', segment)
                    u_match = re.search(r'UserID\.\s*(\S+)', segment)
                    if p_match and o_match and u_match:
                        packing_no = p_match.group(1)
                        order_no = o_match.group(1)
                        user_id = u_match.group(1)
                        if packing_no not in seen_packing_nos:
                            seen_packing_nos.add(packing_no)
                            matched_orders.add(order_no)
                            company_name = "MEDIT EUROPE GMBH" if user_id == "MEDITFRA" else ("MEDIT EUROPE" if user_id == "MEDITRMA" else "")
                            track_no = tracking_map.get(order_no, "")
                            orders.append({"A": "", "B": packing_no, "C": company_name, "D": "", "E": track_no})

        # 누락 감지: 파싱 결과엔 있는데 Picking List에 없는 SAP Order
        missing_orders = []
        for sap in sap_orders:
            if sap not in matched_orders and sap != "Not Found":
                missing_orders.append({
                    "order_no": sap,
                    "tracking": tracking_map.get(sap, "")
                })

        # missing_overrides로 Packing No. 받은 누락 항목 추가
        for sap, packing_no in missing_overrides.items():
            if not packing_no.strip():
                continue
            track_no = tracking_map.get(sap, "")
            # UserID 모르므로 company_name 비움 (필요시 수동 입력 가능)
            orders.append({"A": "", "B": packing_no.strip(), "C": "", "D": "", "E": track_no})

        # 누락 있고 아직 override 안 받은 상태면 → 누락 목록 먼저 반환
        unresolved_missing = [m for m in missing_orders if m['order_no'] not in missing_overrides]
        if unresolved_missing:
            return jsonify({
                "status": "missing",
                "missing": unresolved_missing,
                "message": f"{len(unresolved_missing)}개의 SAP Order가 Picking List에 없습니다."
            }), 200

        if not orders:
            return jsonify({"error": "오더를 찾을 수 없습니다."}), 404

        df = pd.DataFrame(orders)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            worksheet['A1'] = datetime.now().strftime("%Y-%m-%d")
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="picking_list_results.xlsx")
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route('/download_excel', methods=['POST'])
def download_excel():
    data = request.json
    if not data:
        return jsonify({"error": "데이터가 없습니다."}), 400
    df_data = [{"": "", " ": " ", "Tracking": item.get("Tracking Number", ""), "SAP Order(s)": item.get("Ref No", "")} for item in data]
    df = pd.DataFrame(df_data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="extraction_results.xlsx")


@app.route('/print_filter', methods=['POST'])
def print_filter():
    files = request.files.getlist('files')
    target_size = request.form.get('target_size')

    file_data = [
        (f.read(), f.filename)
        for f in files
        if f.filename.lower().endswith('.pdf')
    ]

    writer = PdfWriter()

    def check_size_worker(data_pair):
        data, name = data_pair
        return get_pdf_size(BytesIO(data)) == target_size, data

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        check_results = list(
            executor.map(
                check_size_worker,
                file_data
            )
        )

    grouped = []

    for (is_target, data), (_, name) in zip(
        check_results,
        file_data
    ):
        if not is_target:
            continue

        is_rma_file = bool(
            re.search(
                r'RMA',
                name,
                re.IGNORECASE
            )
        )

        try:
            reader = PdfReader(BytesIO(data))
            total = len(reader.pages)

            if total == 0:
                continue

            first_page = reader.pages[0]

            first_text = (
                first_page.extract_text()
                or ""
            )

            barcode_results = []
            analysis_text = first_text

            if not first_text.strip():
                try:
                    with pdfium_lock:
                        doc = pdfium.PdfDocument(BytesIO(data))
                        page0 = doc[0]
                        bitmap = page0.render(scale=3.0)
                        pil_img = bitmap.to_pil()
                        doc.close()

                    barcode_results = zxingcpp.read_barcodes(
                        pil_img
                    )

                    analysis_text = " ".join(
                        r.text
                        for r in barcode_results
                    )

                except Exception as bc_err:
                    print(
                        f"[PRINT FILTER] "
                        f"Barcode fallback failed for {name}: "
                        f"{bc_err}"
                    )

            # FedEx / DHL 판단
            is_master_fedex = bool(
                re.search(
                    r'##\s*MASTER\s*##',
                    analysis_text,
                    re.IGNORECASE
                )
            )

            fedex_piece_match = re.search(
                r'\b1\s+of\s+(\d+)\b',
                analysis_text,
                re.IGNORECASE
            )

            has_pdf417 = any(
                str(r.format) == "PDF417"
                for r in barcode_results
            )

            fedex_barcode_piece_match = None

            if barcode_results:
                fedex_barcode_piece_match = re.search(
                    r'\b(\d+)\s*/\s*(\d+)\b',
                    analysis_text
                )

            is_fedex_text = bool(
                re.search(
                    r'FEDEX|TRK#|MPS#',
                    analysis_text,
                    re.IGNORECASE
                )
            )

            is_fedex = (
                (
                    is_master_fedex
                    and fedex_piece_match is not None
                )
                or is_fedex_text
                or (
                    bool(barcode_results)
                    and has_pdf417
                )
            )

            # FedEx REF
            ref_no = None

            if is_fedex:
                filename_base = os.path.splitext(name)[0]

                file_ref_match = re.search(
                    r'(?<!\d)(400\d{7})(?!\d)',
                    filename_base
                )

                if file_ref_match:
                    ref_no = file_ref_match.group(1)

            fedex_pages_debug = (
                fedex_piece_match.group(1)
                if fedex_piece_match
                else None
            )

            print(
                f"[PRINT FILTER] {name} | "
                f"FedEx={is_fedex} | "
                f"MASTER={is_master_fedex} | "
                f"FedExPages={fedex_pages_debug} | "
                f"REF={ref_no}"
            )

            print(
                f"[ANALYSIS TEXT] {name} "
                f"(source={'barcode' if barcode_results else 'pdf_text'})\n"
                f"{analysis_text[:3000]}"
            )

            # 페이지 수
            piece_total = None

            if is_fedex:

                if fedex_piece_match:
                    piece_total = int(
                        fedex_piece_match.group(1)
                    )

                elif fedex_barcode_piece_match:
                    piece_total = int(
                        fedex_barcode_piece_match.group(2)
                    )

                else:
                    piece_total = 1

            else:

                dhl_piece_match = re.search(
                    r'\b1\s*/\s*(\d+)\b',
                    analysis_text
                )

                if dhl_piece_match:
                    piece_total = int(
                        dhl_piece_match.group(1)
                    )

            # 기본 페이지 선택
            if piece_total is not None:

                piece_total = min(
                    piece_total,
                    total
                )

                # 정상적인 경우에는 기존 방식 그대로 사용
                pages_to_add = reader.pages[:piece_total]

                # PDF 전체 페이지 수와 바코드상의 페이지 수가
                # 다른 경우에만 전체 페이지 barcode scan
                if piece_total != total:

                    print(
                        f"[PRINT FILTER] "
                        f"Page count mismatch: {name} | "
                        f"PDF={total} | "
                        f"Barcode={piece_total}"
                    )

                    page_pieces = []

                    try:
                        with pdfium_lock:
                            doc = pdfium.PdfDocument(
                                BytesIO(data)
                            )

                            for page_idx in range(total):

                                try:
                                    page = doc[page_idx]

                                    bitmap = page.render(
                                        scale=3.0,
                                        rotation=0
                                    )

                                    pil_img = bitmap.to_pil()

                                    page_barcodes = (
                                        zxingcpp.read_barcodes(
                                            pil_img
                                        )
                                    )

                                    found_piece = None

                                    for barcode in page_barcodes:

                                        barcode_text = (
                                            barcode.text or ""
                                        )

                                        match = re.search(
                                            r'\b(\d+)\s*/\s*(\d+)\b',
                                            barcode_text
                                        )

                                        if not match:
                                            match = re.search(
                                                r'\b(\d+)\s+of\s+(\d+)\b',
                                                barcode_text,
                                                re.IGNORECASE
                                            )

                                        if match:
                                            found_piece = {
                                                "page_idx": page_idx,
                                                "piece_no": int(
                                                    match.group(1)
                                                ),
                                                "total": int(
                                                    match.group(2)
                                                )
                                            }
                                            break

                                    if found_piece:
                                        if (
                                            found_piece["total"]
                                            == piece_total
                                        ):
                                            page_pieces.append(
                                                found_piece
                                            )

                                except Exception as page_err:
                                    print(
                                        f"[PRINT FILTER] "
                                        f"Page barcode error "
                                        f"{name} "
                                        f"(page={page_idx + 1}): "
                                        f"{page_err}"
                                    )

                            doc.close()

                    except Exception as scan_err:
                        print(
                            f"[PRINT FILTER] "
                            f"Full barcode scan failed "
                            f"for {name}: {scan_err}"
                        )

                    page_pieces.sort(
                        key=lambda x: x["piece_no"]
                    )

                    # piece 번호가 실제로 모두 발견된 경우에만
                    # 전체 스캔 결과를 사용
                    piece_numbers = {
                        item["piece_no"]
                        for item in page_pieces
                    }

                    expected_piece_numbers = set(
                        range(
                            1,
                            piece_total + 1
                        )
                    )

                    if (
                        piece_numbers
                        == expected_piece_numbers
                    ):
                        pages_to_add = [
                            reader.pages[
                                item["page_idx"]
                            ]
                            for item in page_pieces
                        ]

                        print(
                            f"[PRINT FILTER] "
                            f"Selected pages for {name}: "
                            f"{[item['page_idx'] + 1 for item in page_pieces]}"
                        )

                    else:
                        print(
                            f"[PRINT FILTER] "
                            f"Could not identify all pieces "
                            f"for {name}. "
                            f"Found={sorted(piece_numbers)}, "
                            f"Expected={sorted(expected_piece_numbers)}"
                        )

            else:
                pages_to_add = reader.pages[:]

            if not pages_to_add:
                continue

            # FedEx REF 표시
            if (
                is_fedex
                and ref_no
                and target_size == 'label'
            ):

                try:

                    for page in pages_to_add:

                        page_width = float(
                            page.mediabox.width
                        )

                        page_height = float(
                            page.mediabox.height
                        )

                        overlay_buf = BytesIO()

                        c = rl_canvas.Canvas(
                            overlay_buf,
                            pagesize=(
                                page_width,
                                page_height
                            )
                        )

                        c.setFillColorRGB(
                            0,
                            0,
                            0
                        )

                        c.setFont(
                            "Helvetica-Bold",
                            14
                        )

                        ref_x = page_width - 120
                        ref_y = page_height - 110

                        c.drawString(
                            ref_x,
                            ref_y,
                            ref_no
                        )

                        c.save()

                        overlay_buf.seek(0)

                        overlay_reader = PdfReader(
                            overlay_buf
                        )

                        page.merge_page(
                            overlay_reader.pages[0]
                        )

                    print(
                        f"[PRINT FILTER] "
                        f"FedEx REF added to "
                        f"{len(pages_to_add)} page(s): "
                        f"{name} -> {ref_no}"
                    )

                except Exception as e:

                    print(
                        f"[PRINT FILTER] "
                        f"REF overlay error "
                        f"{name}: {e}"
                    )

            carrier = (
                "fedex"
                if is_fedex
                else "dhl"
            )

            order = CATEGORY_ORDER.get(
                (
                    carrier,
                    is_rma_file
                ),
                99
            )

            grouped.append(
                (
                    order,
                    pages_to_add
                )
            )

        except Exception as e:

            print(
                f"[PRINT FILTER] "
                f"Error processing {name}: {e}"
            )

            continue

    if not grouped:
        return jsonify({
            "error": (
                f"인쇄 가능한 "
                f"{target_size} 규격의 "
                f"페이지가 없습니다."
            )
        }), 404

    grouped.sort(
        key=lambda g: g[0]
    )

    for _, pages_to_add in grouped:
        for page in pages_to_add:
            writer.add_page(page)

    output = BytesIO()

    writer.write(output)

    output.seek(0)

    return send_file(
        output,
        mimetype='application/pdf'
    )



SPECIAL_CODES = ['3A0113417C0', '3A0112485C0', '3A0113147C0', '3A0113418C0', '3A0113419C0', '3A0112340C0', '6A0111588C0', '5M0111852W0', '5M0112350E0', '3A0113412C0', '3A0113411C0', '3A0111853C0']

@app.route('/generate_delivery_note', methods=['POST'])
def generate_delivery_note():
    source_file = request.files.get('source_file')
    pallet_count = request.form.get('pallet_count', 1)
    box_count = request.form.get('box_count', 0)
    if not source_file:
        return jsonify({"error": "원본 엑셀 파일(20260612.xls)을 업로드해주세요."}), 400
    try:
        pallet_count = int(pallet_count)
        box_count = int(box_count)
    except:
        pallet_count, box_count = 1, 0
    template_path = 'MEDIT_Delivery Note_template.xlsx'
    if not os.path.exists(template_path):
        return jsonify({"error": "템플릿 파일이 서버에 존재하지 않습니다."}), 500
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from copy import copy
        source_content = source_file.read()
        source_df = None
        try:
            source_df = pd.read_excel(BytesIO(source_content))
        except:
            try:
                import xlrd
                book = xlrd.open_workbook(file_contents=source_content, ignore_workbook_corruption=True)
                sheet = book.sheet_by_index(0)
                data = [sheet.row_values(r) for r in range(sheet.nrows)]
                source_df = pd.DataFrame(data[1:], columns=data[0])
            except:
                try: source_df = pd.read_excel(BytesIO(source_content), engine='openpyxl')
                except:
                    dfs = pd.read_html(BytesIO(source_content))
                    if dfs: source_df = dfs[0]
        if source_df is None:
            return jsonify({"error": "엑셀 파일을 읽을 수 없습니다."}), 400

        def format_trkno(val):
            s = str(val).strip()
            digits = re.sub(r'\D', '', s)
            if len(digits) > 12:
                # FedEx처럼 긴 트래킹 번호는 뒤 12자리만 남기고 FED 접두사
                return 'FED' + digits[-12:]
            return s

        if 'TRKNO' in source_df.columns:
            source_df['TRKNO'] = source_df['TRKNO'].apply(format_trkno)

        data_to_fill = source_df[['TRKNO', 'ORDERNO', 'CUSITEMCODE', 'ITEMDETAIL', 'SRL_LOT']]
        special_match_count = sum(1 for code in source_df['CUSITEMCODE'].astype(str).str.strip() if code in SPECIAL_CODES) if 'CUSITEMCODE' in source_df.columns else 0
        total_cartons = int(pd.to_numeric(source_df['BOXCNT'], errors='coerce').sum()) if 'BOXCNT' in source_df.columns else 0
        wb = load_workbook(template_path)
        ws = wb['Sheet1']
        ws['A11'] = f"Carton : {total_cartons} stk"
        ws['B11'] = f"Pallet : {pallet_count} stk"
        num_new_rows = len(data_to_fill)
        sample_row_height = ws.row_dimensions[14].height
        if num_new_rows > 1:
            ws.insert_rows(15, amount=num_new_rows - 1)
        def copy_style(src, dst):
            if src.has_style:
                dst.font, dst.border, dst.fill, dst.number_format, dst.protection, dst.alignment = copy(src.font), copy(src.border), copy(src.fill), copy(src.number_format), copy(src.protection), copy(src.alignment)
        for idx, (_, row) in enumerate(data_to_fill.iterrows()):
            curr = 14 + idx
            if sample_row_height: ws.row_dimensions[curr].height = sample_row_height
            for c_idx, col in enumerate(['TRKNO', 'ORDERNO', 'CUSITEMCODE', 'ITEMDETAIL', 'SRL_LOT'], 1):
                cell = ws.cell(row=curr, column=c_idx, value=row[col])
                copy_style(ws.cell(row=14, column=c_idx), cell)
            for c_idx in [7, 8, 11, 12]:
                src_c = ws.cell(row=14, column=c_idx)
                if curr > 14:
                    dst_c = ws.cell(row=curr, column=c_idx)
                    if src_c.data_type == 'f': dst_c.value = src_c.value.replace('14', str(curr))
                    copy_style(src_c, dst_c)
        # Determine last valid data row and footer row
        last_d = 14 + num_new_rows
        for r in range(ws.max_row, 14, -1):
            if any(ws.cell(row=r, column=c).value for c in range(1, 6)): last_d = r; break
        foot_r = last_d + 2

        # Find the exact signature rows after insertion by scanning for "DATEN :"
        sig_start_row = None
        sig_end_row = None
        for r in range(14, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val == "DATEN :":
                sig_start_row = r
            if val == "Unterschrift :":
                sig_end_row = r
                break

        if sig_start_row and sig_end_row:
            # Calculate A4 printable height dynamically based on margins (A4 height is 841.68 points)
            top_margin = ws.page_margins.top if ws.page_margins.top is not None else 0.75
            bottom_margin = ws.page_margins.bottom if ws.page_margins.bottom is not None else 1.0
            a4_page_height = 841.68 - (top_margin * 72.0) - (bottom_margin * 72.0) - 5.0
            default_h = ws.sheet_format.defaultRowHeight or 12.75

            # Map each row (up to foot_r) to its calculated page number
            page_map = {}
            current_page = 1
            current_height = 0.0
            for r in range(1, foot_r + 1):
                h = ws.row_dimensions[r].height or default_h
                if current_height + h > a4_page_height:
                    current_page += 1
                    current_height = h
                else:
                    current_height += h
                page_map[r] = current_page

            # If signature block spans pages, insert a manual break before it
            manual_break_row = None
            if page_map.get(sig_start_row, 1) != page_map.get(foot_r, 1):
                manual_break_row = sig_start_row - 1

            # Re-simulate with manual break
            actual_page_map = {}
            current_page = 1
            current_height = 0.0
            for r in range(1, foot_r + 1):
                h = ws.row_dimensions[r].height or default_h
                if manual_break_row and r == manual_break_row + 1:
                    current_page += 1
                    current_height = h
                elif current_height + h > a4_page_height:
                    current_page += 1
                    current_height = h
                else:
                    current_height += h
                actual_page_map[r] = current_page

            # Insert manual page breaks at page transitions
            break_points = [r for r in range(2, foot_r + 1)
                            if actual_page_map.get(r) != actual_page_map.get(r - 1)]
            from openpyxl.worksheet.pagebreak import Break
            for bp in reversed(break_points):
                ws.row_breaks.append(Break(id=bp - 1))

        # Set footer: &G inserts the VML-linked image in the center of the page footer
        ws.oddFooter.center.text = "&G"

        # Logo image (header area, sheet drawing)
        logo_p = 'image1.png'
        try: from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, XDRPositiveSize2D
        except: from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker; from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        if os.path.exists(logo_p):
            img = XLImage(logo_p)
            img.anchor = OneCellAnchor(_from=AnchorMarker(col=0, colOff=pixels_to_EMU(60), row=3, rowOff=100), ext=XDRPositiveSize2D(cx=pixels_to_EMU(310), cy=pixels_to_EMU(160)))
            ws.add_image(img)

        ws.page_setup.paperSize, ws.page_setup.orientation, ws.print_area = 9, 'portrait', f'A1:E{foot_r + 5}'
        # Save workbook to buffer first
        output = BytesIO()
        wb.save(output)

        # Inject VML footer image into the saved xlsx (zip) directly from the template.
        # openpyxl does not support header/footer images natively, so we patch the zip.
        foot_p = 'image2.png'
        if os.path.exists(foot_p) and os.path.exists(template_path):
            # Read required VML content from template
            with zipfile.ZipFile(template_path, 'r') as tmpl_z:
                vml_content = tmpl_z.read('xl/drawings/vmlDrawing1.vml')
                vml_rels_content = tmpl_z.read('xl/drawings/_rels/vmlDrawing1.vml.rels')
                footer_img_content = tmpl_z.read('xl/media/image2.png')

            # Re-write the xlsx zip with VML entries added / patched
            VML_CT = '<Default Extension="vml" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/>'
            VML_REL_ENTRY = (
                '<Relationship Id="rIdVML" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" '
                'Target="../drawings/vmlDrawing1.vml"/>'
            )
            output.seek(0)
            original_bytes = output.read()
            patched = BytesIO()
            with zipfile.ZipFile(BytesIO(original_bytes), 'r') as src_z, \
                 zipfile.ZipFile(patched, 'w', compression=zipfile.ZIP_DEFLATED) as dst_z:
                existing = {i.filename for i in src_z.infolist()}
                for item in src_z.infolist():
                    data = src_z.read(item.filename)
                    if item.filename == '[Content_Types].xml':
                        # Add VML content type if missing
                        txt = data.decode('utf-8', errors='replace')
                        if 'vmlDrawing' not in txt:
                            txt = txt.replace('</Types>', VML_CT + '</Types>')
                        data = txt.encode('utf-8')
                    elif item.filename == 'xl/worksheets/_rels/sheet1.xml.rels':
                        # Add legacyDrawingHF (VML) relationship
                        txt = data.decode('utf-8', errors='replace')
                        if 'vmlDrawing' not in txt:
                            txt = txt.replace('</Relationships>', VML_REL_ENTRY + '</Relationships>')
                        data = txt.encode('utf-8')
                    elif item.filename == 'xl/worksheets/sheet1.xml':
                        # Add <legacyDrawingHF> before </worksheet>.
                        # openpyxl's root <worksheet> element does NOT declare xmlns:r globally;
                        # r: only appears inline on the <drawing> tag. To avoid "undeclared prefix"
                        # errors, we include xmlns:r explicitly on legacyDrawingHF itself.
                        txt = data.decode('utf-8', errors='replace')
                        if 'legacyDrawingHF' not in txt:
                            R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                            legacy_tag = f'<legacyDrawingHF xmlns:r="{R_NS}" r:id="rIdVML"/>'
                            txt = txt.replace('</worksheet>', legacy_tag + '</worksheet>')
                        data = txt.encode('utf-8')
                    elif item.filename == 'xl/media/image2.png':
                        # Replace with template's footer image
                        data = footer_img_content
                    dst_z.writestr(item, data)
                # Add VML files if not already present
                if 'xl/drawings/vmlDrawing1.vml' not in existing:
                    dst_z.writestr('xl/drawings/vmlDrawing1.vml', vml_content)
                if 'xl/drawings/_rels/vmlDrawing1.vml.rels' not in existing:
                    dst_z.writestr('xl/drawings/_rels/vmlDrawing1.vml.rels', vml_rels_content)
                if 'xl/media/image2.png' not in existing:
                    dst_z.writestr('xl/media/image2.png', footer_img_content)
            patched.seek(0)
            output = patched
        else:
            output.seek(0)

        res = send_file(output, as_attachment=True, download_name="MEDIT_Delivery Note.xlsx")
        res.headers.update({'X-Special-Match-Count': str(special_match_count), 'X-Total-Cartons': str(total_cartons), 'X-Pallet-Count': str(pallet_count), 'X-Box-Count': str(box_count)})
        return res
    except Exception as e:
        return jsonify({"error": str(e)}), 500

CATEGORY_ORDER_BY_LABEL = {category_label(carrier, is_rma): order for (carrier, is_rma), order in CATEGORY_ORDER.items()}

@app.route('/highlight_picking_list', methods=['POST'])
def highlight_picking_list():
    file = request.files.get('picking_list')
    tracking_data_json = request.form.get('tracking_data', '[]')
    if not file:
        return jsonify({"error": "파일을 업로드해주세요."}), 400
    try:
        import pdfplumber
        from pypdf import PdfReader, PdfWriter
        import reportlab.pdfgen.canvas as rl_canvas
        from reportlab.lib.colors import HexColor
        import io

        # Ref No -> Tracking Number / Category 매핑
        tracking_map = {}
        category_map = {}
        try:
            tracking_list = json.loads(tracking_data_json)
            for item in tracking_list:
                ref = item.get('Ref No', '').strip()
                track = item.get('Tracking Number', '').strip()
                if ref and track and track != "Not Found":
                    tracking_map[ref] = track
                if ref:
                    category_map[ref] = item.get('Category', 'normal-dhl')
        except:
            pass

        FEDEX_TRACKING_MIN_LEN = 15

        def normalize_order_no(s):
            if not s:
                return None
            return re.sub(r'^RMA-?', '', s.strip(), flags=re.IGNORECASE)

        def is_fedex_order(order_no):
            track = tracking_map.get(order_no, "")
            digits = re.sub(r'\D', '', track)
            return len(digits) >= FEDEX_TRACKING_MIN_LEN

        input_bytes = file.read()
        reader = PdfReader(io.BytesIO(input_bytes))
        writer = PdfWriter()

        highlight_color = HexColor('#FFD700')
        border_color = HexColor('#FF8C00')
        star_color = HexColor('#CC0000')
        fedex_color = HexColor("#F73B0C")

        fixed_pages = []      # 요약 페이지 등, 순서 그대로 유지할 페이지 (원래 인덱스, page)
        order_pages = []      # 개별 오더 페이지 (카테고리 순서, 원래 인덱스, page) - 정렬 대상

        with pdfplumber.open(io.BytesIO(input_bytes)) as pdf:
            for page_num in range(len(reader.pages)):
                page = reader.pages[page_num]
                plumber_page = pdf.pages[page_num]
                text = page.extract_text() or ""

                if "Packing No." not in text or "OrderNo." not in text:
                    fixed_pages.append((page_num, page))
                    continue

                matched_codes = [code for code in SPECIAL_CODES if code in text]
                words = plumber_page.extract_words()

                page_order_no = None
                page_user_id = None

                order_match = re.search(r'OrderNo\.\s*(\S+)', text)
                if order_match:
                    page_order_no = order_match.group(1)

                userid_match = re.search(r'UserID\.\s*(\S+)', text)
                if userid_match:
                    page_user_id = userid_match.group(1)

                fedex_lines = []
                for w in words:
                    if w['text'] != 'OrderNo.':
                        continue
                    line_top = w['top']
                    line_words = sorted(
                        [lw for lw in words if abs(lw['top'] - line_top) < 2],
                        key=lambda lw: lw['x0']
                    )
                    try:
                        pos = next(i for i, lw in enumerate(line_words) if lw['text'] == 'OrderNo.')
                    except StopIteration:
                        continue
                    if pos + 1 >= len(line_words):
                        continue
                    order_value_word = line_words[pos + 1]
                    order_no = order_value_word['text']

                    if not is_fedex_order(order_no):
                        continue

                    userid_word = next((lw for lw in line_words if lw['text'] == 'UserID.'), None)
                    x_start = order_value_word['x1']
                    x_end = userid_word['x0'] if userid_word else x_start + 80
                    fedex_lines.append((line_top, x_start, x_end))

                page_width = float(page.mediabox.width)
                page_height = float(page.mediabox.height)

                if matched_codes or fedex_lines:
                    overlay_buf = io.BytesIO()
                    c = rl_canvas.Canvas(overlay_buf, pagesize=(page_width, page_height))
                    drawn_lines = set()

                    for code in matched_codes:
                        for w in words:
                            if w['text'] == code:
                                top = w['top']
                                line_key = round(top, 1)
                                if line_key in drawn_lines:
                                    continue
                                drawn_lines.add(line_key)

                                line_words = [lw for lw in words if abs(lw['top'] - top) < 2]
                                if not line_words:
                                    continue

                                x0 = min(lw['x0'] for lw in line_words) - 4
                                x1 = max(lw['x1'] for lw in line_words) + 4
                                line_top = min(lw['top'] for lw in line_words)
                                line_bottom = max(lw['bottom'] for lw in line_words)

                                box_y0 = page_height - line_bottom - 2
                                box_y1 = page_height - line_top + 2
                                box_height = box_y1 - box_y0

                                c.setFillColorRGB(1, 0.84, 0, alpha=0.35)
                                c.setStrokeColor(border_color)
                                c.setLineWidth(1)
                                c.rect(x0, box_y0, x1 - x0 + 20, box_height,
                                       fill=1, stroke=1)

                                c.setFillColor(star_color)
                                c.setFont("Helvetica-Bold", 11)
                                c.drawString(max(x0 - 18, 2), box_y0 + 2, "★")

                    for line_top, x_start, x_end in fedex_lines:
                        mid_x = (x_start + x_end) / 2
                        font_size = 15
                        y_offset = 17
                        y = page_height - line_top - y_offset
                        c.setFillColor(fedex_color)
                        c.setFont("Helvetica-Bold", font_size)
                        text_width = c.stringWidth("FEDEX", "Helvetica-Bold", font_size)
                        c.drawString(mid_x - text_width / 2, y, "FEDEX")

                    c.save()
                    overlay_buf.seek(0)
                    overlay_reader = PdfReader(overlay_buf)
                    overlay_page = overlay_reader.pages[0]
                    page.merge_page(overlay_page)

                # 정렬 순서 결정: tracking_data의 Category 우선, 없으면 FedEx 판정으로 추정
                category = category_map.get(normalize_order_no(page_order_no))
                if category is None:
                    category = 'normal-fedex' if fedex_lines else 'normal-dhl'
                order = CATEGORY_ORDER_BY_LABEL.get(category, 99)




                # ---------------------------------------------------------
                # Category 결정
                # ---------------------------------------------------------

                # UserID로 RMA / Normal 판단
                user_id_upper = (page_user_id or "").strip().upper()

                if "RMA" in user_id_upper:
                    is_rma = True
                elif "FRA" in user_id_upper:
                    is_rma = False
                else:
                    # 혹시 UserID를 못 읽은 경우 기존 tracking_data의 Category 사용
                    existing_category = category_map.get(
                        normalize_order_no(page_order_no)
                    )

                    if existing_category:
                        is_rma = existing_category.startswith("rma-")
                    else:
                        is_rma = False


                # DHL / FedEx 판단
                if fedex_lines:
                    carrier = "fedex"
                else:
                    # tracking_data에 있는 Tracking Number를 fallback으로 사용
                    track = tracking_map.get(
                        normalize_order_no(page_order_no),
                        ""
                    )

                    digits = re.sub(r'\D', '', track)

                    if len(digits) >= FEDEX_TRACKING_MIN_LEN:
                        carrier = "fedex"
                    else:
                        carrier = "dhl"


                # 최종 Category
                category = category_label(carrier, is_rma)

                order = CATEGORY_ORDER_BY_LABEL.get(category, 99)

                order_pages.append((order, page_num, page))

        # 카테고리 순서로 정렬 (같은 카테고리 내에서는 원래 순서 유지 - stable sort)
        order_pages.sort(key=lambda t: (t[0], t[1]))

        # 요약 페이지는 그대로, 개별 오더 페이지는 정렬된 순서로
        for _, page in fixed_pages:
            writer.add_page(page)
        for _, _, page in order_pages:
            writer.add_page(page)

        output = io.BytesIO()
        writer.write(output)
        output.seek(0)
        return send_file(output, mimetype='application/pdf',
                          as_attachment=False,
                          download_name='picking_highlighted.pdf')
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000)
