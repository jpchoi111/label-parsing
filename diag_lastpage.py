import zipfile
from openpyxl import load_workbook

# Check the test2 file structure
fname = 'MEDIT_Delivery Note3_footer_test2.xlsx'
wb = load_workbook(fname)
ws = wb.active

print('Max row:', ws.max_row)
print('Print area:', ws.print_area)
print('Row breaks:', sorted([b.id for b in ws.row_breaks.brk]))
print()

# Find last row with content
last_content = 0
for r in range(ws.max_row, 1, -1):
    if any(ws.cell(row=r, column=c).value for c in range(1, 13)):
        last_content = r
        break

print('Last row with content:', last_content)

# Show rows around the end
print()
print('Last 15 rows:')
for r in range(max(1, ws.max_row - 14), ws.max_row + 1):
    h = ws.row_dimensions[r].height
    vals = [ws.cell(row=r, column=c).value for c in range(1, 4)]
    has_val = any(vals)
    print(f'  Row {r:3d} h={str(h):6} {"<<DATA" if has_val else ""}  {[str(v)[:20] if v else "" for v in vals]}')
